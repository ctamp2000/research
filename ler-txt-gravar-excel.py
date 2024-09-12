'''Ler arquivo txt e grava xlsx'''
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
#arquivo = open('acm-temp.txt','r')  # nome do arquivo texto , modo de abertura;
arquivo = open('input.txt','r')  # input.txt is your input file , r means that you will read this input file.
c1 = sheet.cell(row=1, column=1) # column titles (in the first line): author, title, and abstract (in this example the output file has 3 columns).
c1.value = 'author'
c1 = sheet.cell(row=1, column=2)
c1.value = 'title'
c1 = sheet.cell(row=1, column=3)
c1.value = 'abstract'
planilha_linha = 2
for linha in arquivo:   # read your entire input file, 1 line at a time.
    valor = linha
    posicao = linha.find('author')  # search for 'author'
    if posicao != -1: # found 'author'
        c1 = sheet.cell(row=planilha_linha, column=1)  # author
        c1.value = linha[posicao+10:linha.__len__()]  # name
 #       print(planilha_linha,'author',linha[posicao+9:linha.__len__()])
    else:
        posicao = linha.find('title')
        if posicao != -1: # found 'title'
            posicao = linha.find('ktitle')
            if posicao == -1: # it is not booktitle
                c1 = sheet.cell(row=planilha_linha, column=2)  # title
                c1.value = linha[posicao+10:linha.__len__()]  # name
 #               print('title',linha[posicao+8:linha.__len__()])
        else:
            posicao = linha.find('abstract')
            if posicao != -1: # found 'abstract'
                c1 = sheet.cell(row=planilha_linha, column=3)  # abstract
                c1.value = linha[posicao+12:linha.__len__()]  # name
 #               print('abstract',linha[posicao+11:linha.__len__()])
                planilha_linha += 1
#    print('for', planilha_linha)
wb.save('C:\\Users\\ctamp\\Downloads\\acm.xlxs') # put the path and name of your output file here

